"""Planungsgenauigkeit: Plan vs. IST-Umsatz je Zeitebene und Aggregationsebene."""
import streamlit as st
from pathlib import Path
import sys
sys.path.insert(0, str(Path(__file__).parent.parent.parent))

from ui.session import get_conn, get_gmbh, require_db, get_budgetjahr
import pandas as pd
import io

require_db()
conn = get_conn()
gmbh = get_gmbh()
planjahr = get_budgetjahr()

st.title("Planungsgenauigkeit")

plan_rows = conn.execute(
    "SELECT * FROM planung WHERE CAST(strftime('%Y', datum) AS INTEGER)=? ORDER BY fil_nr, datum",
    (planjahr,),
).fetchall()

if not plan_rows:
    st.info(
        f"Noch keine Planungsdaten für {planjahr} vorhanden. "
        "Bitte zuerst unter **Planung ausführen** eine Berechnung starten."
    )
    st.stop()

# IST-Daten für das Budgetjahr (soweit vorhanden)
ist_rows = conn.execute(
    "SELECT fil_nr, datum, umsatz FROM ist_umsatz WHERE datum LIKE ?",
    (f"{planjahr}-%",),
).fetchall()
ist_lookup = {(str(r["fil_nr"]), r["datum"]): r["umsatz"] for r in ist_rows}
has_ist = len(ist_lookup) > 0

def _g(r, k, default=0.0):
    try:
        v = r[k]
        return v if v is not None else default
    except (IndexError, KeyError):
        return default

df = pd.DataFrame([{
    "fil_nr":     str(r["fil_nr"]),
    "datum":      r["datum"],
    "wochentag":  r["wochentag"],
    "bundesland": _g(r, "bundesland", "") or "",
    "IST Basis":  _g(r, "ist_vj"),
    "Budget":     _g(r, "budget") or _g(r, "tagesumsatz_plan") or _g(r, "gesamt_plan"),
    "IST aktuell": ist_lookup.get((str(r["fil_nr"]), r["datum"])),
} for r in plan_rows])

df["datum_dt"] = pd.to_datetime(df["datum"])

# Bundesland aus filialen-Tabelle als Fallback
if df["bundesland"].isna().all() or (df["bundesland"] == "").all():
    bl_map = {r[0]: r[1] for r in conn.execute("SELECT fil_nr, bundesland FROM filialen").fetchall()}
    df["bundesland"] = df["fil_nr"].map(bl_map).fillna("?")

# ── Steuerung ──────────────────────────────────────────────────────────────
cf1, cf2 = st.columns(2)
with cf1:
    fil_filter = st.multiselect(
        "Filtern auf Filiale(n) (leer = alle)",
        sorted(df["fil_nr"].unique()),
        placeholder="Filialen auswählen...",
        key="plangenau_fil_filter",
    )
with cf2:
    bl_filter = st.multiselect(
        "Filtern auf Bundesland (leer = alle)",
        sorted(df["bundesland"].dropna().unique()),
        placeholder="Bundesland auswählen...",
        key="plangenau_bl_filter",
    )

if fil_filter:
    df = df[df["fil_nr"].isin(fil_filter)]
if bl_filter:
    df = df[df["bundesland"].isin(bl_filter)]

c1, c2 = st.columns(2)
with c1:
    zeit_ebene = st.selectbox("Zeit-Ebene", ["Tag", "Woche", "Monat", "Jahr"], index=2,
                               key="plangenau_zeit")
with c2:
    entity_ebene = st.selectbox("Aggregations-Ebene", ["Filiale", "Bundesland", "Gesamt"],
                                 key="plangenau_entity")

# ── Zeit-Gruppierung ────────────────────────────────────────────────────────
MON = ["Jan", "Feb", "Mär", "Apr", "Mai", "Jun", "Jul", "Aug", "Sep", "Okt", "Nov", "Dez"]

if zeit_ebene == "Tag":
    df["Zeit"] = df["datum_dt"].dt.strftime("%d.%m.%Y")
    df["_sort"] = df["datum_dt"]
elif zeit_ebene == "Woche":
    iso = df["datum_dt"].dt.isocalendar()
    df["Zeit"] = "KW " + iso["week"].astype(str).str.zfill(2) + "/" + iso["year"].astype(str)
    df["_sort"] = df["datum_dt"].dt.to_period("W").apply(lambda p: p.start_time)
elif zeit_ebene == "Monat":
    df["Zeit"] = df["datum_dt"].dt.month.map(lambda m: MON[m - 1]) + " " + df["datum_dt"].dt.year.astype(str)
    df["_sort"] = df["datum_dt"].dt.to_period("M").apply(lambda p: p.start_time)
else:
    df["Zeit"] = df["datum_dt"].dt.year.astype(str)
    df["_sort"] = df["datum_dt"].dt.year

group_keys = ["Zeit", "_sort"]
if entity_ebene == "Filiale":
    group_keys = ["fil_nr"] + group_keys
elif entity_ebene == "Bundesland":
    group_keys = ["bundesland"] + group_keys

# Budget nur für Tage zählen, an denen IST-Umsatz bereits importiert ist —
# sonst ist die Abweichung in angebrochenen Monaten/Wochen irreführend.
df["_budget_ist"] = df["Budget"].where(df["IST aktuell"].notna())

agg = (df.groupby([k for k in group_keys if k != "_sort"], as_index=False)
       .agg({
           "IST Basis":    "sum",
           "Budget":       "sum",
           "_budget_ist":  lambda x: x.sum() if x.notna().any() else None,
           "IST aktuell":  lambda x: x.sum() if x.notna().any() else None,
           "_sort":        "min",
       })
       .sort_values([k for k in (["_sort"] if entity_ebene == "Gesamt"
                                  else [group_keys[0], "_sort"])]))

agg["Abw. €"] = agg.apply(
    lambda x: round(float(x["IST aktuell"]) - float(x["_budget_ist"]), 2)
    if not pd.isna(x["IST aktuell"]) and not pd.isna(x["_budget_ist"]) else None,
    axis=1,
)
agg["Abw. %"] = agg.apply(
    lambda x: round(float(x["Abw. €"]) / float(x["_budget_ist"]) * 100, 1)
    if not pd.isna(x["Abw. €"]) and not pd.isna(x["_budget_ist"])
    and float(x["_budget_ist"]) != 0 else None,
    axis=1,
)

rename = {"fil_nr": "Filiale", "bundesland": "Bundesland"}
disp = agg.drop(columns=["_sort", "_budget_ist"]).rename(columns=rename)
lead = [c for c in ["Filiale", "Bundesland", "Zeit"] if c in disp.columns]
ordered = lead + ["IST Basis", "Budget", "IST aktuell", "Abw. €", "Abw. %"]
disp = disp[[c for c in ordered if c in disp.columns]]

# ── Kennzahlen ─────────────────────────────────────────────────────────────
def _de(val) -> str:
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return "–"
    return f"{float(val):,.0f}".replace(",", "X").replace(".", ",").replace("X", ".")

tot_basis  = agg["IST Basis"].sum()
tot_bud    = agg["Budget"].sum()
tot_bud_ist = agg["_budget_ist"].sum() if agg["_budget_ist"].notna().any() else None
tot_ist    = agg["IST aktuell"].sum() if has_ist and agg["IST aktuell"].notna().any() else None

ist_bis = df.loc[df["IST aktuell"].notna(), "datum_dt"].max() if has_ist else None

m1, m2, m3, m4 = st.columns(4)
m1.metric("IST Basis", f"{_de(tot_basis)} €")
m2.metric("Budget", f"{_de(tot_bud)} €")
if tot_ist is not None and tot_bud_ist is not None:
    abw_e = tot_ist - tot_bud_ist
    abw_p = abw_e / tot_bud_ist * 100 if tot_bud_ist != 0 else 0.0
    m3.metric("IST aktuell", f"{_de(tot_ist)} €")
    m4.metric("Abw. IST/Budget", f"{'+' if abw_e >= 0 else ''}{_de(abw_e)} € ({abw_p:+.1f} %)",
              help=f"Verglichen wird nur das Budget der Tage mit importiertem IST-Umsatz "
                   f"(Budget bis dahin: {_de(tot_bud_ist)} €).")
else:
    m3.metric("IST aktuell", "– (noch kein Import)")
    m4.metric("Abw. IST/Budget", "–")

if ist_bis is not None:
    st.caption(
        f"IST-Umsätze importiert bis **{ist_bis.strftime('%d.%m.%Y')}**. "
        "Abweichungen (€/%) vergleichen IST nur mit dem Budget bis zu diesem Datum — "
        "angebrochene Wochen/Monate werden anteilig gerechnet."
    )

st.divider()

# ── Tabelle ─────────────────────────────────────────────────────────────────
def _fmt_de(val):
    try:
        if pd.isna(val):
            return ""
    except (TypeError, ValueError):
        pass
    try:
        return f"{float(val):,.0f}".replace(",", "X").replace(".", ",").replace("X", ".")
    except (TypeError, ValueError):
        return ""

def _fmt_pct(val):
    try:
        if pd.isna(val):
            return ""
    except (TypeError, ValueError):
        pass
    try:
        return f"{float(val):+.1f} %"
    except (TypeError, ValueError):
        return ""

num_cols = ["IST Basis", "Budget", "IST aktuell", "Abw. €"]

# Pre-format numeric columns directly (avoids Pandas Styler cell-limit issues)
disp_fmt = disp.copy()
for c in num_cols:
    if c in disp_fmt.columns:
        disp_fmt[c] = disp_fmt[c].apply(_fmt_de)
if "Abw. %" in disp_fmt.columns:
    disp_fmt["Abw. %"] = disp_fmt["Abw. %"].apply(_fmt_pct)

st.dataframe(disp_fmt, use_container_width=True, hide_index=True, height=560)

# ── Excel-Export ────────────────────────────────────────────────────────────
st.divider()
buf = io.BytesIO()
with pd.ExcelWriter(buf, engine="openpyxl") as writer:
    disp.to_excel(writer, index=False, sheet_name="Planungsgenauigkeit")
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter
    ws = writer.sheets["Planungsgenauigkeit"]
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for col_idx in range(1, len(disp.columns) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 16

suffix = f"_{entity_ebene}_{zeit_ebene}"
st.download_button(
    label="📥 Excel herunterladen",
    data=buf.getvalue(),
    file_name=f"Planungsgenauigkeit_{planjahr}{suffix}_{gmbh.replace(' ', '_')}.xlsx",
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    type="primary",
)
